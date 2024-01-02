
# excel_functions.py
# file where all excel Read and Write functions are there

from openpyxl import load_workbook

class Excel_funtion:

    # fetch the row count
    @staticmethod
    def row_count(file, sheet_name):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    # fetch the column count
    @staticmethod
    def column_count(file, sheet_name):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_column

    # read data from excel file
    @staticmethod
    def read_data(file, sheet_name, row_number, column_number):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_number, column=column_number).value

    # write data into excel file
    @staticmethod
    def write_data(file, sheet_name, row_number, column_number, data):
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(file)

    def read_data_value(self, file, sheet_name):
        final_list = []
        workbook = load_workbook(file)
        sheet = workbook[sheet_name]
        for r in range(2, self.row_count(file, sheet_name)):
            row_list = []
            for c in range(6, self.column_count(file, sheet_name) + 1):
                row_list.append(sheet.cell(row=r, column=c).value)
            final_list.append(row_list)
        return final_list




